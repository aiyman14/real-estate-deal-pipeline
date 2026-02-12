"""
Tests for Excel writer, including append functionality.
"""

import pytest
import tempfile
from pathlib import Path

from src.render.excel_writer import (
    write_excel,
    append_to_excel,
    get_sheet_name_for_country,
    SHEET_DEAL_LIST,
    SHEET_SWEDEN,
    SHEET_DENMARK,
    SHEET_FINLAND,
)


class TestSheetNameMapping:
    """Test country to sheet name mapping."""

    def test_sweden_sheet(self):
        assert get_sheet_name_for_country("Sweden") == SHEET_SWEDEN

    def test_denmark_sheet(self):
        assert get_sheet_name_for_country("Denmark") == SHEET_DENMARK

    def test_finland_sheet(self):
        assert get_sheet_name_for_country("Finland") == SHEET_FINLAND

    def test_lowercase_sweden(self):
        assert get_sheet_name_for_country("sweden") == SHEET_SWEDEN

    def test_unknown_defaults_to_sweden(self):
        assert get_sheet_name_for_country("Norway") == SHEET_SWEDEN

    def test_empty_defaults_to_sweden(self):
        assert get_sheet_name_for_country("") == SHEET_SWEDEN

    def test_none_defaults_to_sweden(self):
        assert get_sheet_name_for_country(None) == SHEET_SWEDEN


class TestAppendToExcel:
    """Test append_to_excel functionality."""

    def test_append_creates_file_if_not_exists(self, tmp_path):
        """Test that append creates a new file if it doesn't exist."""
        output_file = tmp_path / "test.xlsx"
        columns = ["Name", "Value"]
        row = {"Name": "Test", "Value": "123"}

        row_num = append_to_excel(row, columns, output_file, "Sheet1")

        assert output_file.exists()
        assert row_num == 2  # First data row after header

    def test_append_adds_to_existing_file(self, tmp_path):
        """Test that append adds rows to existing file."""
        output_file = tmp_path / "test.xlsx"
        columns = ["Name", "Value"]

        # First row
        row1 = {"Name": "First", "Value": "1"}
        row_num1 = append_to_excel(row1, columns, output_file, "Sheet1")
        assert row_num1 == 2

        # Second row
        row2 = {"Name": "Second", "Value": "2"}
        row_num2 = append_to_excel(row2, columns, output_file, "Sheet1")
        assert row_num2 == 3

        # Third row
        row3 = {"Name": "Third", "Value": "3"}
        row_num3 = append_to_excel(row3, columns, output_file, "Sheet1")
        assert row_num3 == 4

    def test_append_to_different_sheets(self, tmp_path):
        """Test that append can add to different sheets in same file."""
        output_file = tmp_path / "deals.xlsx"
        columns = ["Country", "Price"]

        # Add to Sweden sheet
        sweden_row = {"Country": "Sweden", "Price": "100"}
        sweden_row_num = append_to_excel(sweden_row, columns, output_file, SHEET_SWEDEN)
        assert sweden_row_num == 2

        # Add to Denmark sheet
        denmark_row = {"Country": "Denmark", "Price": "200"}
        denmark_row_num = append_to_excel(denmark_row, columns, output_file, SHEET_DENMARK)
        assert denmark_row_num == 2

        # Add another to Sweden
        sweden_row2 = {"Country": "Sweden", "Price": "150"}
        sweden_row_num2 = append_to_excel(sweden_row2, columns, output_file, SHEET_SWEDEN)
        assert sweden_row_num2 == 3  # Second row in Sweden sheet

    def test_append_preserves_existing_sheets(self, tmp_path):
        """Test that appending to one sheet doesn't affect others."""
        from openpyxl import load_workbook

        output_file = tmp_path / "deals.xlsx"
        columns = ["Name"]

        # Add to Sheet1
        append_to_excel({"Name": "A"}, columns, output_file, "Sheet1")
        append_to_excel({"Name": "B"}, columns, output_file, "Sheet1")

        # Add to Sheet2
        append_to_excel({"Name": "X"}, columns, output_file, "Sheet2")

        # Verify both sheets exist with correct data
        wb = load_workbook(output_file)
        assert "Sheet1" in wb.sheetnames
        assert "Sheet2" in wb.sheetnames

        # Sheet1 should have 3 rows (header + 2 data)
        assert wb["Sheet1"].max_row == 3
        # Sheet2 should have 2 rows (header + 1 data)
        assert wb["Sheet2"].max_row == 2


class TestSheetConstants:
    """Test sheet name constants are correct."""

    def test_deal_list_name(self):
        assert SHEET_DEAL_LIST == "Deal list"

    def test_sweden_name(self):
        assert SHEET_SWEDEN == "Sweden"

    def test_denmark_name(self):
        assert SHEET_DENMARK == "Denmark"

    def test_finland_name(self):
        assert SHEET_FINLAND == "Finland"
