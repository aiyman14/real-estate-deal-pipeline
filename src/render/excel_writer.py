"""
Excel (.xlsx) output writer.

Supports both creating new files and appending to existing files with multiple sheets.
"""

from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None
    load_workbook = None


# Sheet names for the 4-sheet structure Viktor uses
SHEET_DEAL_LIST = "Deal list"
SHEET_SWEDEN = "Sweden"
SHEET_DENMARK = "Denmark"
SHEET_FINLAND = "Finland"


def get_sheet_name_for_country(country: str) -> str:
    """Map country to sheet name for transactions."""
    country_lower = country.lower() if country else "sweden"
    if country_lower == "denmark":
        return SHEET_DENMARK
    elif country_lower == "finland":
        return SHEET_FINLAND
    else:
        return SHEET_SWEDEN


def write_excel(
    rows: List[Dict[str, Any]],
    columns: List[str],
    output_path: Path,
    sheet_name: str = "Deals",
) -> None:
    """
    Write rows to an Excel file (creates new file).

    Args:
        rows: List of row dicts (already rendered/formatted)
        columns: Column names in order
        output_path: Path to .xlsx file
        sheet_name: Name of the worksheet
    """
    if Workbook is None:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header row with styling
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            value = row.get(col_name, "")
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths (approximate)
    for col_idx, col_name in enumerate(columns, start=1):
        max_length = len(col_name)
        for row in rows:
            val = str(row.get(col_name, ""))
            if len(val) > max_length:
                max_length = min(len(val), 50)  # Cap at 50
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_length + 2

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def append_to_excel(
    row: Dict[str, Any],
    columns: List[str],
    output_path: Path,
    sheet_name: str,
) -> int:
    """
    Append a single row to an Excel file, creating file/sheet if needed.

    Args:
        row: Row dict (already rendered/formatted)
        columns: Column names in order
        output_path: Path to .xlsx file
        sheet_name: Name of the worksheet to append to

    Returns:
        Row number where the data was written
    """
    if Workbook is None:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Load existing workbook or create new one
    if output_path.exists():
        wb = load_workbook(output_path)
    else:
        wb = Workbook()
        # Remove default "Sheet" if we're creating a fresh workbook
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # Get or create the sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Check if sheet has headers
        has_headers = ws.cell(row=1, column=1).value is not None
    else:
        ws = wb.create_sheet(sheet_name)
        has_headers = False

    # Write headers if sheet is empty
    if not has_headers:
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
        # Auto-adjust column widths for headers
        for col_idx, col_name in enumerate(columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = min(len(col_name) + 2, 50)

    # Find next empty row
    next_row = ws.max_row + 1
    if not has_headers:
        next_row = 2  # First data row after headers

    # Write the row
    for col_idx, col_name in enumerate(columns, start=1):
        value = row.get(col_name, "")
        ws.cell(row=next_row, column=col_idx, value=value)

    # Save
    wb.save(output_path)

    return next_row


def init_deals_workbook(
    output_path: Path,
    inbound_columns: List[str],
    sweden_columns: List[str],
    denmark_columns: List[str],
    finland_columns: List[str],
) -> None:
    """
    Initialize a new workbook with all 4 sheets pre-created with headers.

    Args:
        output_path: Path to .xlsx file
        inbound_columns: Columns for Deal list sheet
        sweden_columns: Columns for Sweden sheet
        denmark_columns: Columns for Denmark sheet
        finland_columns: Columns for Finland sheet
    """
    if Workbook is None:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Remove default "Sheet"
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Create each sheet with headers
    sheets_config = [
        (SHEET_DEAL_LIST, inbound_columns),
        (SHEET_SWEDEN, sweden_columns),
        (SHEET_DENMARK, denmark_columns),
        (SHEET_FINLAND, finland_columns),
    ]

    for sheet_name, columns in sheets_config:
        ws = wb.create_sheet(sheet_name)
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            ws.column_dimensions[get_column_letter(col_idx)].width = min(len(col_name) + 2, 50)

    wb.save(output_path)
